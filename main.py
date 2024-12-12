from typing import Tuple
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

ctk.set_appearance_mode('dark')
ctk.set_default_color_theme('dark-blue')

class telaInicial(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.config()
        self.aparencia()
        self.loginFrame()
        self.corpo()

    def config(self):
        self.geometry('600x500')
        self.title('Tela de login')

    def mudarAparencia(self, novaAparencia):
        ctk.set_appearance_mode(novaAparencia)

    def aparencia(self):
        labelTema = ctk.CTkLabel(self, text='Tema', font=('Arvo bold', 14), text_color=['#000', '#fff'])
        labelTema.place(x=10, y=430)

        tema = ctk.CTkComboBox(self, values=['Dark', 'Light'], command=self.mudarAparencia)
        tema.place(x=10, y=460)

    def loginFrame(self):
        frame = ctk.CTkFrame(self, height=120, width=600, fg_color='#22394a')
        frame.place(x=0, y=0)

        labelFrame = ctk.CTkLabel(frame, text='Tela de Login', font=('Arvo bold', 24), text_color='#fff')
        labelFrame.place(x=230, y=50)

    def corpo(self):
        labelLogin = ctk.CTkLabel(self, text='Login:', font=('Arvo bold', 18), text_color=['#000', '#fff'])
        labelLogin.place(x=280, y=150) 

        entryLogin = ctk.CTkEntry(self, width=300, height=35, font=('Arvo bold', 12), placeholder_text='Nome de Usuário')
        entryLogin.place(x=165, y=190)

        labelSenha = ctk.CTkLabel(self, text='Senha:', font=('Arvo bold', 18), text_color=['#000', '#fff'])
        labelSenha.place(x=280, y=250) 

        entrySenha = ctk.CTkEntry(self, width=300, height=35, font=('Arvo bold', 12), placeholder_text='Sua Senha')
        entrySenha.place(x=165, y=290)

        def entrar():
            nomeUsuario = entryLogin.get()
            senhaUsuario = entrySenha.get()

            print(nomeUsuario, senhaUsuario)
            

            if nomeUsuario == 'fabiano' and senhaUsuario == '161318':
                cadastro = ctk.CTkToplevel(inicial)

                #config
                cadastro.geometry('700x600')
                cadastro.title('Cadastro Computadores')
                cadastro.resizable(width=False, height=False)

                #frame
                cadastroFrame = ctk.CTkFrame(cadastro, height=150, width=700, fg_color='#22394a')
                cadastroFrame.place(x=0, y=0)

                cadastroTitulo = ctk.CTkLabel(cadastroFrame, text='Cadastro de Computadores', font=('Arvo bold', 24), text_color='#fff')
                cadastroTitulo.place(x=200, y=70)

                #Aparencia

                cadastroTemaLabel = ctk.CTkLabel(cadastro, text='Tema', font=('Arvo bold', 14))
                cadastroTemaLabel.place(x=10, y=530)

                def cadastroMudarTema(cadastroNovoTema):
                    ctk.set_appearance_mode(cadastroNovoTema)

                cadastroTema = ctk.CTkComboBox(cadastro, values=['Dark', 'Light'], command=cadastroMudarTema)
                cadastroTema.place(x=10, y=560)

                #Corpo

                labelNome = ctk.CTkLabel(cadastro, text='Nome do Computador:', font=('Arvo bold', 14))
                labelNome.place(x=10, y=170)

                entryNome = ctk.CTkEntry(cadastro, placeholder_text='Digite Aqui:', height=40, width=250)
                entryNome.place(x=10, y=200)


                labelPrecoPago = ctk.CTkLabel(cadastro, text='Preço Pago:', font=('Arvo bold', 14))
                labelPrecoPago.place(x=10, y=260)

                entryPrecoPago = ctk.CTkEntry(cadastro, placeholder_text='Digite Aqui:', height=40, width=250)
                entryPrecoPago.place(x=10, y=290)


                labelGasto = ctk.CTkLabel(cadastro, text='Gastos com o Computador:', font=('Arvo bold', 14))
                labelGasto.place(x=400, y=170)

                entryGasto = ctk.CTkEntry(cadastro, placeholder_text='Digite Aqui:', height=40, width=250)
                entryGasto.place(x=400, y=200)


                labelVendido = ctk.CTkLabel(cadastro, text='Preço de Venda:', font=('Arvo bold', 14))
                labelVendido.place(x=400, y=260)

                entryVendido = ctk.CTkEntry(cadastro, placeholder_text='Digite Aqui:', height=40, width=250)
                entryVendido.place(x=400, y=290)


                labelObs = ctk.CTkLabel(cadastro, text='Observação:', font=('Arvo bold', 14))
                labelObs.place(x=10, y=350)

                entryObs = ctk.CTkTextbox(cadastro, height=70, width=300, border_width=2)
                entryObs.place(x=10, y=380)



                ficheiro = pathlib.Path('Cadastro.xlsx')
                if ficheiro.exists():
                    pass
                else:
                    ficheiro = Workbook()
                    folha = ficheiro.active

                    folha['A1'] = 'Nome do Computador'
                    folha['B1'] = 'Preço'
                    folha['C1'] = 'Gastos'
                    folha['D1'] = 'Preço de venda'
                    folha['E1'] = 'Lucro'
                    folha['F1'] = 'Obs'

                    ficheiro.save('Cadastro.xlsx')
                    


                #Botões
                def enviar():
                    nome = entryNome.get()
                    preco = float(entryPrecoPago.get())
                    gasto = float(entryGasto.get())
                    vendido = float(entryVendido.get())
                    obs = entryObs.get(0.0, END)

                    lucro = vendido - (preco + gasto)

                    if nome == '' or preco == '' or gasto == '' or vendido == '':
                        messagebox.showerror('Sistema da Nina', 'Fabiano,\nVocê precisa preencher pelo menos um campo!')
                    else:

                        ficheiro = openpyxl.load_workbook('Cadastro.xlsx')

                        folha = ficheiro.active

                        folha.cell(column=1, row=folha.max_row+1, value=nome)
                        folha.cell(column=2, row=folha.max_row, value=preco)
                        folha.cell(column=3, row=folha.max_row, value=gasto)
                        folha.cell(column=4, row=folha.max_row, value=vendido)
                        folha.cell(column=5, row=folha.max_row, value=lucro)
                        folha.cell(column=6, row=folha.max_row, value=obs)

                        ficheiro.save(r'Cadastro.xlsx')

                    messagebox.showinfo('Sistema da Nina', 'Fabiano, Você salvou os dados!')

                
                def limpar():
                    entryNome.delete(0, END)
                    entryPrecoPago.delete(0, END)
                    entryGasto.delete(0, END)
                    entryVendido.delete(0, END)
                    entryObs.delete(0.0, END)

                    messagebox.showinfo('Sistema da Nina', 'Fabiano, Você limpou a tela!')
                    



                botaoEnviar = ctk.CTkButton(cadastro, text='Enviar', command=enviar)
                botaoEnviar.place(x=200, y=560)

                botaoLimpar = ctk.CTkButton(cadastro, text='Limpar', command=limpar)
                botaoLimpar.place(x=400, y=560)

            
            elif nomeUsuario != 'fabiano':
                erro = messagebox.showwarning('Sistema da Nina', 'Fabiano,\nNome de Usuário INCORRETO')
            
            elif senhaUsuario != '161318':
                erro1 = messagebox.showwarning('Sistema da Nina', 'Fabiano, \nSenha INCORRETA')

            elif senhaUsuario != '161318' and nomeUsuario != 'fabiano':
                erro2 = messagebox.showwarning('Sistema da Nina', 'Senha e Nome de Usuário Incorretos')


        botaoEntrar = ctk.CTkButton(self, text='Entrar', command=entrar)
        botaoEntrar.place(x=240, y=350)



    
inicial = telaInicial()
inicial.mainloop()
